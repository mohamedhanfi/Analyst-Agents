"""Unit tests for shared/reader.FileReader (no CrewAI involved)."""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from shared.core.reader import FileReader


@pytest.fixture
def reader(tmp_path):
    run_dir = tmp_path / "run123"
    return FileReader(run_dir)


def test_read_csv_comma(tmp_path, reader):
    path = tmp_path / "data.csv"
    pd.DataFrame({"a": [1, 2, 3], "b": ["x", "y", "z"]}).to_csv(path, index=False)
    r = reader.read(path)
    assert r.file_type == "csv"
    assert r.row_count == 3
    assert r.column_count == 2
    assert r.separator == ","
    assert r.extracted_path is not None
    assert Path(r.extracted_path).is_file()
    assert r.file_hash.startswith("sha256:")


def test_read_csv_semicolon(tmp_path, reader):
    path = tmp_path / "data.csv"
    path.write_text("a;b\n1;2\n3;4\n", encoding="utf-8")
    r = reader.read(path)
    assert r.separator == ";"
    assert r.row_count == 2


def test_read_xlsx_discover_only(tmp_path, reader):
    path = tmp_path / "multi.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sales"
    ws1.append(["a", "b"])
    for i in range(10):
        ws1.append([i, i + 1])
    ws2 = wb.create_sheet("Meta")
    ws2.append(["k"])
    for i in range(5):
        ws2.append([i])
    wb.save(path)

    r = reader.read(path)
    assert r.file_type == "xlsx"
    assert r.extracted_path is None  # nothing extracted yet
    assert [s.name for s in r.sheets] == ["Sales", "Meta"]
    assert r.sheets[0].row_count == 11  # header + 10
    assert r.sheets[1].row_count == 6


def test_extract_sheet_writes_only_selected(tmp_path, reader):
    path = tmp_path / "multi.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sales"
    ws1.append(["a", "b"])
    for i in range(3):
        ws1.append([i, i + 1])
    ws2 = wb.create_sheet("Meta")
    ws2.append(["k"])
    ws2.append([1])
    wb.save(path)

    out = reader.extract_sheet(path, "Sales")
    assert "multi__Sales.csv" in out
    assert Path(out).is_file()
    extracted_files = list(reader.extracted_dir.glob("*.csv"))
    assert len(extracted_files) == 1  # only the selected one
    assert extracted_files[0].name == "multi__Sales.csv"


def test_safe_filename(tmp_path, reader):
    path = tmp_path / "t.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales👑Data"
    ws.append(["a"])
    ws.append([1])
    wb.save(path)
    out = reader.extract_sheet(path, "Sales👑Data")
    assert "t__Sales_Data.csv" in out


def test_unsupported_extension(tmp_path, reader):
    path = tmp_path / "notes.txt"
    path.write_text("hello")
    with pytest.raises(ValueError):
        reader.read(path)


def test_hash_is_stable(tmp_path, reader):
    path = tmp_path / "d.csv"
    path.write_bytes(b"a,b\n1,2\n")
    h1 = reader.read(path).file_hash
    h2 = reader.read(path).file_hash
    assert h1 == h2 and h1.startswith("sha256:")