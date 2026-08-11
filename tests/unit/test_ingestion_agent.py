"""Unit tests for agents/ingestion_agent.run_ingestion (deterministic path).

No API key required: use_crew=False exercises the full orchestration through
the shared/core classes, covering the §2.1 deliverables: extracted CSV +
data_profile.json + business_context.json under a run dir, multi-sheet sheet
selection, Generic Mode fallback, PII redaction and failure handling.
"""
from __future__ import annotations

import json

import openpyxl
import pandas as pd
import pytest

from agents.ingestion_agent import run_ingestion
from shared.utils import load_config


@pytest.fixture
def cfg():
    return load_config(require_key=False)


@pytest.fixture
def sales_csv(tmp_path):
    path = tmp_path / "sales.csv"
    pd.DataFrame({
        "date": ["2024-01-01", "2024-01-02", "2024-01-03", "2024-01-04",
                 "2024-01-05", "2024-01-06", "2024-01-07"],
        "product": ["A", "B", "A", "B", "A", "B", "A"],
        "category": ["W", "W", "G", "G", "W", "G", "W"],
        "revenue": [100.0, 150.0, 200.0, 90.0, 120.0, 180.0, 140.0],
        "quantity": [5, 7, 3, 9, 6, 4, 8],
    }).to_csv(path, index=False)
    return path


def _multi_sheet_xlsx(tmp_path, name="multi.xlsx"):
    path = tmp_path / name
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sales"
    ws1.append(["date", "product", "revenue"])
    for i in range(1, 15):
        ws1.append([f"2024-01-{i:02d}", f"P{i}", i * 10])
    ws2 = wb.create_sheet("Meta")
    ws2.append(["k"])
    for i in range(4):
        ws2.append([i])
    wb.save(path)
    return path


def _business_answers():
    return iter(["Track revenue", "sales", "Top products?", "Set targets"])


def _xlsx_answers():
    return iter(["Sales", "Track revenue", "sales", "Top products?",
                 "Set targets"])


def _provider(answers):
    return lambda _prompt: next(answers)


# ---------------------------------------------------------------------------
# CSV happy path
# ---------------------------------------------------------------------------


def test_csv_produces_all_artifacts(tmp_path, cfg, sales_csv):
    s = run_ingestion(str(sales_csv), run_dir=tmp_path / "r1", cfg=cfg,
                      answer_provider=_provider(_business_answers()))
    assert s["status"] == "passed"
    assert s["row_count"] == 7
    assert s["column_count"] == 5
    assert s["generic_mode"] is False

    assert (tmp_path / "r1" / "data" / "extracted" / "sales.csv").is_file()
    profile = json.loads(
        (tmp_path / "r1" / "metadata" / "data_profile.json")
        .read_text(encoding="utf-8"))
    assert profile["row_count"] == 7
    assert profile["validation_status"] == "passed"

    ctx = json.loads(
        (tmp_path / "r1" / "knowledge" / "business_context.json")
        .read_text(encoding="utf-8"))
    assert ctx["goal_summary"] == "Track revenue"
    assert ctx["business_questions"] == ["Top products?"]
    assert (tmp_path / "r1" / "logs" / "run.jsonl").is_file()


def test_csv_logs_stage_lifecycle(tmp_path, cfg, sales_csv):
    s = run_ingestion(str(sales_csv), run_dir=tmp_path / "r2", cfg=cfg,
                      answer_provider=_provider(_business_answers()))
    lines = [json.loads(l) for l in
             open(s["log_path"], encoding="utf-8").read().splitlines()]
    kinds = [l["kind"] for l in lines]
    assert kinds[0] == "stage_start"
    assert kinds[-1] == "stage_end"
    assert "tool_call" in kinds
    assert all(l.get("run_id") for l in lines)


# ---------------------------------------------------------------------------
# XLSX paths
# ---------------------------------------------------------------------------


def test_xlsx_single_sheet(tmp_path, cfg):
    path = tmp_path / "single.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["a", "b"])
    for i in range(6):
        ws.append([i, i + 1])
    wb.save(path)

    s = run_ingestion(str(path), run_dir=tmp_path / "r3", cfg=cfg,
                      answer_provider=_provider(_business_answers()))
    assert s["status"] == "passed"
    assert s["sheet_used"] == "Data"
    assert s["row_count"] == 6
    assert "single__Data.csv" in s["extracted_path"]


def test_xlsx_multi_sheet_uses_chosen(tmp_path, cfg):
    path = _multi_sheet_xlsx(tmp_path)
    s = run_ingestion(str(path), run_dir=tmp_path / "r4", cfg=cfg,
                      answer_provider=_provider(_xlsx_answers()))
    assert s["status"] == "passed"
    assert s["sheet_used"] == "Sales"
    assert s["generic_mode"] is False
    assert s["row_count"] == 14
    assert "multi__Sales.csv" in s["extracted_path"]


def test_xlsx_multi_sheet_timeout_uses_largest(tmp_path, cfg):
    path = _multi_sheet_xlsx(tmp_path)

    def slow(_prompt):
        import time
        time.sleep(2)
        return "late"

    s = run_ingestion(str(path), run_dir=tmp_path / "r5", cfg=cfg,
                      answer_provider=slow, timeout_seconds=0.2)
    assert s["status"] == "passed"
    assert s["generic_mode"] is True
    assert s["context_confidence"] == 0.0
    assert s["sheet_used"] == "Sales"
    assert "multi__Sales.csv" in s["extracted_path"]

    ctx = json.loads(
        (tmp_path / "r5" / "knowledge" / "business_context.json")
        .read_text(encoding="utf-8"))
    assert ctx["generic_mode"] is True


# ---------------------------------------------------------------------------
# PII redaction
# ---------------------------------------------------------------------------


def test_profile_redacts_pii(tmp_path, cfg):
    path = tmp_path / "pii.csv"
    pd.DataFrame({
        "email": ["a@x.com", "b@x.com", "c@x.com", "d@x.com", "e@x.com"],
        "revenue": [10, 20, 30, 40, 50],
    }).to_csv(path, index=False)
    s = run_ingestion(str(path), run_dir=tmp_path / "r6", cfg=cfg,
                      answer_provider=_provider(_business_answers()))
    assert s["pii_columns"] == ["email"]
    profile = json.loads(
        (tmp_path / "r6" / "metadata" / "data_profile.json")
        .read_text(encoding="utf-8"))
    for row in profile["sample"]:
        assert row["email"] == "[REDACTED]"


# ---------------------------------------------------------------------------
# Failures
# ---------------------------------------------------------------------------


def test_invalid_file_returns_failed_summary(tmp_path, cfg):
    path = tmp_path / "notes.txt"
    path.write_text("hello,world\n1,2\n", encoding="utf-8")
    s = run_ingestion(str(path), run_dir=tmp_path / "r7", cfg=cfg,
                      answer_provider=_provider(_business_answers()))
    assert s["status"] == "failed"
    assert s["validation_status"] == "failed"
    assert any(e["code"] == "unsupported_format" for e in s["errors"])
    assert not (tmp_path / "r7" / "metadata" / "data_profile.json").exists()


def test_too_few_rows_fails(tmp_path, cfg):
    path = tmp_path / "small.csv"
    path.write_text("a,b\n1,2\n", encoding="utf-8")
    s = run_ingestion(str(path), run_dir=tmp_path / "r8", cfg=cfg,
                      answer_provider=_provider(_business_answers()))
    assert s["status"] == "failed"
    assert any(e["code"] == "min_rows" for e in s["errors"])

