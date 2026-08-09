"""Unit tests for shared/validation.FileValidator (no CrewAI involved)."""
from __future__ import annotations

import openpyxl
import pandas as pd
import pytest

from shared.core.validation import FileValidator


@pytest.fixture
def validator():
    return FileValidator(max_file_size_mb=200.0, min_rows=5)


def _write_bytes(tmp_path, name, data: bytes):
    p = tmp_path / name
    p.write_bytes(data)
    return str(p)


def test_valid_csv_passes(tmp_path, validator):
    df = pd.DataFrame({"a": [1, 2, 3, 4, 5, 6], "b": ["x"] * 6})
    path = tmp_path / "ok.csv"
    df.to_csv(path, index=False)
    r = validator.validate(str(path))
    assert r.validation_status == "passed"
    assert r.extension_ok and r.size_ok and r.signature_ok
    assert r.parser_ok and r.rows_ok
    assert r.row_count == 6
    assert r.errors == []


def test_csv_with_bom_passes(tmp_path, validator):
    path = tmp_path / "bom.csv"
    path.write_bytes(b"\xef\xbb\xbfa,b\n1,2\n3,4\n5,6\n7,8\n9,10\n", )
    r = validator.validate(str(path))
    assert r.validation_status == "passed"


def test_less_than_min_rows_fails(tmp_path, validator):
    path = tmp_path / "tiny.csv"
    path.write_text("a,b\n1,2\n3,4\n", encoding="utf-8")
    r = validator.validate(str(path))
    assert r.validation_status == "failed"
    assert r.errors[0].code == "min_rows"
    assert r.row_count == 2


def test_unsupported_extension_fails(tmp_path, validator):
    (tmp_path / "notes.txt").write_text("hello", encoding="utf-8")
    r = validator.validate(str(tmp_path / "notes.txt"))
    assert r.validation_status == "failed"
    assert r.errors[0].code == "unsupported_format"


def test_oversized_file_fails(tmp_path):
    small = FileValidator(max_file_size_mb=0.000001, min_rows=5)
    path = tmp_path / "big.csv"
    path.write_bytes(b"a,b\n" + b"1,2\n" * 6)
    r = small.validate(str(path))
    assert r.validation_status == "failed"
    assert r.errors[0].code == "file_too_large"


def test_nul_bytes_csv_fails_signature(tmp_path, validator):
    r = validator.validate(_write_bytes(tmp_path, "bad.csv", b"a,b\x00\n1,2\n"))
    assert r.validation_status == "failed"
    assert r.errors[0].code == "invalid_signature"


def test_xlsx_content_renamed_csv_fails_signature(tmp_path, validator):
    r = validator.validate(_write_bytes(tmp_path, "fake.csv", b"PK\x00\x04junkjunk"))
    assert r.validation_status == "failed"
    assert r.errors[0].code == "invalid_signature"


def test_corrupt_xlsx_fails_parser(tmp_path, validator):
    r = validator.validate(_write_bytes(tmp_path, "bad.xlsx", b"PK\x03\x04not-a-real-zip"))
    assert r.validation_status == "failed"
    assert r.errors[0].code == "parser_error"


def test_valid_xlsx_passes(tmp_path, validator):
    path = tmp_path / "ok.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["a", "b"])
    for i in range(10):
        ws.append([i, i * 2])
    wb.save(path)
    r = validator.validate(str(path))
    assert r.validation_status == "passed"
    assert r.row_count == 11  # header + 10 data rows


def test_xlsx_uses_largest_sheet_not_sum(tmp_path, validator):
    path = tmp_path / "multi.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "small"
    ws1.append(["a"])
    ws1.append(["1"])
    ws2 = wb.create_sheet("big")
    ws2.append(["a"])
    for i in range(20):
        ws2.append([i])
    wb.save(path)
    r = validator.validate(str(path))
    assert r.validation_status == "passed"
    assert r.row_count == 21  # not 22 = sum of both sheets