"""FileReader — reads, discovers and extracts, never returns raw data.

Separation of concerns (multi-agent friendly):
- FileValidator  -> "is this file valid?"  (shared/validation.py)
- FileReader     -> "what is inside, how do we read it?" (this module)
- Agent          -> "which sheet / data do we keep?" (decides between them)

CSV has a single sheet, so extraction is immediate.
XLSX is discovery-only in read(); the agent picks a sheet, then
extract_sheet() streams that one sheet to CSV (read_only, row by row).
"""
from __future__ import annotations

import csv
import hashlib
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

from shared.schemas import ReadResult, SheetInfo

SUPPORTED_EXTENSIONS = {".csv", ".xlsx"}
CSV_SEPARATORS = (",", ";")
CSV_ENCODINGS = ("utf-8-sig", "latin-1")
HASH_CHUNK = 1024 * 1024


class FileReader:
    def __init__(self, run_dir: str | Path):
        self.run_dir = Path(run_dir)
        self.extracted_dir = self.run_dir / "data" / "extracted"
        self.extracted_dir.mkdir(parents=True, exist_ok=True)

    # ------------------------------------------------------------------ API

    def read(self, path: str | Path, extension: str | None = None) -> ReadResult:
        path = Path(path)
        extension = (extension or path.suffix).lower()
        if not extension.startswith("."):
            extension = f".{extension}"
        if extension not in SUPPORTED_EXTENSIONS:
            raise ValueError(f"Unsupported extension: {extension}")
        if not path.is_file():
            raise FileNotFoundError(f"Not a file: {path}")

        file_hash = self._calculate_hash(path)

        if extension == ".csv":
            return self._read_csv(path, file_hash)
        return self._read_xlsx(path, file_hash)

    def extract_sheet(self, path: str | Path, sheet_name: str) -> str:
        """Stream one XLSX sheet to CSV (read_only, no full DataFrame)."""
        path = Path(path)
        safe_name = self._safe_filename(sheet_name)
        output_path = self.extracted_dir / f"{path.stem}__{safe_name}.csv"

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb[sheet_name]
            with open(output_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow(row)
        finally:
            wb.close()
        return str(output_path)

    # ------------------------------------------------------------- internals

    def _read_csv(self, path: Path, file_hash: str) -> ReadResult:
        separator = self._detect_csv_separator(path)

        df = None
        encoding_used = None
        last_error: Exception | None = None
        for encoding in CSV_ENCODINGS:
            try:
                df = pd.read_csv(path, sep=separator, encoding=encoding)
                encoding_used = encoding
                break
            except Exception as exc:  # noqa: BLE001 -- try next encoding
                last_error = exc

        if df is None:
            raise ValueError(f"Unable to parse CSV file: {last_error}")

        output_path = self.extracted_dir / f"{path.stem}.csv"
        df.to_csv(output_path, index=False, encoding="utf-8")

        return ReadResult(
            file_type="csv",
            file_name=path.name,
            file_hash=file_hash,
            row_count=len(df),
            column_count=len(df.columns),
            separator=separator,
            encoding=encoding_used,
            extracted_path=str(output_path),
        )

    def _detect_csv_separator(self, path: Path) -> str:
        first_line = ""
        for encoding in CSV_ENCODINGS:
            try:
                with open(path, "r", encoding=encoding, newline="") as f:
                    first_line = f.readline()
                break
            except UnicodeDecodeError:
                continue

        if not first_line:
            raise ValueError("Unable to read the first line of the CSV.")

        comma_count = first_line.count(",")
        semicolon_count = first_line.count(";")
        if semicolon_count > comma_count:
            return ";"
        return ","

    def _read_xlsx(self, path: Path, file_hash: str) -> ReadResult:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            sheets = []
            for ws in wb.worksheets:
                sheets.append(SheetInfo(name=ws.title, row_count=self._count_rows(ws)))
            return ReadResult(
                file_type="xlsx",
                file_name=path.name,
                file_hash=file_hash,
                sheets=sheets,
            )
        finally:
            wb.close()

    @staticmethod
    def _count_rows(ws) -> int:
        """Count data rows (header included). openpyxl max_row in read-only
        mode can be unreliable, so count explicitly."""
        count = 0
        for _ in ws.iter_rows(values_only=True):
            count += 1
        return count

    def _calculate_hash(self, path: Path) -> str:
        sha256 = hashlib.sha256()
        with open(path, "rb") as f:
            while chunk := f.read(HASH_CHUNK):
                sha256.update(chunk)
        return f"sha256:{sha256.hexdigest()}"

    @staticmethod
    def _safe_filename(name: str) -> str:
        allowed = "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-_"
        return "".join(ch if ch in allowed else "_" for ch in name)